"""
Run Snapshot Packager
"""

import json
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from backend.config import DATA_DIR, RUNS_DIR
from backend.agents.text_to_sql.db_writer import write_metrics_db
from backend.storage.s3_runs import upload_run_directory


def _make_run_dir(run_id: str) -> Path:
    run_dir = RUNS_DIR / run_id
    run_dir.mkdir(parents=True, exist_ok=True)
    return run_dir


def _copy_raw_csvs(run_dir: Path) -> list:
    copied = []
    for fname in ["case4_weekly_submissions.csv", "case4_weekly_premium.csv",
                  "case4_pipeline.csv", "case4_loss_indicators.csv"]:
        src = DATA_DIR / fname
        if src.exists():
            shutil.copy2(src, run_dir / fname)
            copied.append(fname)
    return copied


def _write_merged_xlsx(df: pd.DataFrame, run_dir: Path) -> str:
    path = run_dir / "merged_metrics.xlsx"
    wb   = Workbook()
    ws   = wb.active
    ws.title = "Merged Metrics"

    header_fill = PatternFill("solid", fgColor="1A1A2E")
    header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=10)

    cols   = list(df.columns)
    df_out = df.copy()
    df_out["week_ending"] = df_out["week_ending"].dt.strftime("%Y-%m-%d")

    for ci, col in enumerate(cols, 1):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.fill, cell.font = header_fill, header_font
        cell.alignment = Alignment(horizontal="center")

    for ri, row in enumerate(df_out.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val).font = Font(name="Calibri", size=10)

    for ci in range(1, len(cols) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 22

    wb.save(path)
    return "merged_metrics.xlsx"


def _write_llm1_input(payload: dict, run_dir: Path) -> str:
    (run_dir / "llm1_input_payload.json").write_text(
        json.dumps(payload, indent=2, default=str), encoding="utf-8"
    )
    return "llm1_input_payload.json"


def _write_llm1_output(llm_chain: dict, run_dir: Path) -> str:
    llm1 = llm_chain.get("llm1", {})
    data = {
        "model":       llm1.get("model"),
        "prompt_file": llm1.get("prompt_file"),
        "parsed":      llm1.get("parsed", {}),
        "raw_output":  llm1.get("raw_output", ""),
        "usage":       llm1.get("usage", {}),
    }
    (run_dir / "llm1_output.json").write_text(
        json.dumps(data, indent=2, default=str), encoding="utf-8"
    )
    return "llm1_output.json"


def _write_signals_md(payload: dict, llm_chain: dict, run_dir: Path) -> str:
    path  = run_dir / "signals_and_enrichment.md"
    ps    = payload.get("portfolio_summary", {})
    sc    = payload.get("signal_counts", {})
    lines = [
        "# MosAIc Signal Detection Report\n\n",
        f"**Generated:** {payload.get('generated_at', '')}\n",
        f"**Report Week:** {ps.get('report_week', '')}\n\n---\n\n",
        "## Portfolio Summary\n",
        f"- Weeks: {ps.get('total_weeks_analysed')} ({ps.get('week_range_start')} → {ps.get('week_range_end')})\n",
        f"- YTD Actual GWP: £{ps.get('total_ytd_actual_gwp', 0):,.0f}\n",
        f"- YTD Plan GWP:   £{ps.get('total_ytd_plan_gwp', 0):,.0f}\n",
        f"- vs Plan: {ps.get('portfolio_ytd_gwp_ratio', 0):.1%}\n\n---\n\n",
        "## Signal Counts\n",
        f"- S1: {sc.get('S1_structural_underperformance', 0)} | S2: {sc.get('S2_hit_rate_collapse', 0)} | "
        f"S3: {sc.get('S3_loss_ratio_deterioration', 0)} | S4: {sc.get('S4_profitable_outperformance', 0)}\n",
        f"- Total Anomalies: {sc.get('total_anomalies', 0)}\n\n---\n\n",
        "## All Detected Concerns\n\n",
    ]
    for c in payload.get("all_concerns", []):
        lines += [
            f"### [{c.get('signal_id')}] {c.get('lob')} — {c.get('signal_name')}\n",
            f"- Severity: {c.get('severity')} | Impact: £{c.get('impact_score', 0):,.0f}\n",
        ]
        if c.get("root_cause"):
            lines += [f"- Root Cause: {c.get('root_cause')}\n",
                      f"- Detail: {c.get('root_cause_detail', '')}\n"]
        lines.append("\n")

    lines.append("## All Detected Opportunities\n\n")
    for o in payload.get("all_opportunities", []):
        lines += [
            f"### [{o.get('signal_id')}] {o.get('lob')} — {o.get('signal_name')}\n",
            f"- Health: {o.get('health_verdict')} | Surplus: £{o.get('gwp_surplus', 0):,.0f}\n",
            f"- {o.get('health_note', '')}\n\n",
        ]

    lines += [
        "---\n\n## LLM1 Prioritisation Output\n\n",
        f"```json\n{json.dumps(llm_chain.get('llm1', {}).get('parsed', {}), indent=2)}\n```\n\n",
        f"**Analyst Notes:** {llm_chain.get('analyst_notes', '')}\n",
    ]
    path.write_text("".join(lines), encoding="utf-8")
    return "signals_and_enrichment.md"


def _write_prompts(llm_chain: dict, run_dir: Path) -> list:
    written = []
    for key in ("llm1", "llm2"):
        data    = llm_chain.get(key, {})
        fname   = data.get("prompt_file", f"{key}_prompt.txt")
        content = data.get("prompt_text", "")
        (run_dir / fname).write_text(content, encoding="utf-8")
        written.append(fname)
    return written


def _write_narrative_html(html: str, run_dir: Path) -> str:
    (run_dir / "narrative_report.html").write_text(html, encoding="utf-8")
    return "narrative_report.html"


def _write_dashboard_json(payload: dict, llm_chain: dict, run_dir: Path) -> str:
    data = {
        "portfolio_summary": payload.get("portfolio_summary", {}),
        "lob_snapshot":      payload.get("lob_snapshot", []),
        "weekly_series":     payload.get("weekly_series", []),
        "top_concerns":      llm_chain.get("top_concerns", []),
        "top_opportunity":   llm_chain.get("top_opportunity", {}),
        "all_concerns":      payload.get("all_concerns", []),
        "all_opportunities": payload.get("all_opportunities", []),
        "anomalies":         payload.get("anomalies", {}),
        "pipeline_friction": payload.get("pipeline_friction", []),
        "analyst_notes":     llm_chain.get("analyst_notes", ""),
        "session_id":        llm_chain.get("session_id", ""),
    }
    (run_dir / "dashboard_data.json").write_text(
        json.dumps(data, indent=2, default=str), encoding="utf-8"
    )
    return "dashboard_data.json"


def _write_zip(run_dir: Path, run_id: str) -> str:
    zip_path = run_dir / f"snapshot_{run_id}.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for fpath in run_dir.iterdir():
            if fpath.suffix != ".zip":
                zf.write(fpath, fpath.name)
    return f"snapshot_{run_id}.zip"


def create_snapshot(run_id: str, df: pd.DataFrame, payload: dict, llm_chain: dict) -> dict:
    run_dir  = _make_run_dir(run_id)
    manifest = {
        "run_id":     run_id,
        "run_dir":    str(run_dir),
        "created_at": datetime.utcnow().isoformat() + "Z",
        "files":      {},
    }

    manifest["files"]["raw_csvs"]       = _copy_raw_csvs(run_dir)
    manifest["files"]["merged_xlsx"]    = _write_merged_xlsx(df, run_dir)
    manifest["files"]["metrics_db"]     = write_metrics_db(df, run_dir)   # ← SQLite for agent
    manifest["files"]["llm1_input"]     = _write_llm1_input(payload, run_dir)
    manifest["files"]["llm1_output"]    = _write_llm1_output(llm_chain, run_dir)
    manifest["files"]["signals_md"]     = _write_signals_md(payload, llm_chain, run_dir)
    manifest["files"]["prompts"]        = _write_prompts(llm_chain, run_dir)
    manifest["files"]["narrative_html"] = _write_narrative_html(llm_chain.get("narrative_html", ""), run_dir)
    manifest["files"]["dashboard_json"] = _write_dashboard_json(payload, llm_chain, run_dir)
    manifest["files"]["zip"]            = _write_zip(run_dir, run_id)

    # Everything above writes to local disk exactly as before — this is
    # the one new step. On local dev / Docker Compose (S3_RUNS_BUCKET
    # unset) this is a no-op. On AWS, it uploads the now-complete run
    # directory to S3, since Fargate's container filesystem doesn't
    # survive a restart or redeploy the way local disk does.
    upload_run_directory(run_dir, run_id)

    print(f"[Snapshot] Created: {run_dir}")
    return manifest
